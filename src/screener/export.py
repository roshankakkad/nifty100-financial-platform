from pathlib import Path
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

ROOT = Path(__file__).resolve().parents[2]
sys.path.append(str(ROOT / "src"))
OUTPUT_FILE = ROOT / "output" / "screener_output.xlsx"

from screener.presets import presets, run_preset

EXPORT_COLUMNS = [
    "company_id", "company_name", "broad_sector", "year", "composite_quality_score", "sector_relative_score",
    "return_on_equity_pct", "return_on_capital_employed_pct", "net_profit_margin_pct",
    "operating_profit_margin_pct", "debt_to_equity", "interest_coverage", "free_cash_flow_cr",
    "revenue_cagr_3yr", "revenue_cagr_5yr", "pat_cagr_5yr", "eps_cagr_5yr",
    "asset_turnover", "sales", "net_profit", "pe_ratio", "pb_ratio", "dividend_yield_pct",
    "dividend_payout_ratio_pct", "market_cap_crore"
]

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="F4CCCC")
HEADER = PatternFill("solid", fgColor="1F4E78")
THIN = Side(style="thin", color="D9E2F3")


def _sheet_name(name):
    return name.replace("_", " ").title()[:31]


def _meets(col, value, rules, row):
    if pd.isna(value):
        return False
    checks = {
        "return_on_equity_pct": ("roe_min", ">"),
        "debt_to_equity": ("de_max", "<="),
        "free_cash_flow_cr": ("fcf_min", ">"),
        "revenue_cagr_3yr": ("revenue_cagr_3yr_min", ">"),
        "revenue_cagr_5yr": ("revenue_cagr_5yr_min", ">"),
        "pat_cagr_5yr": ("pat_cagr_5yr_min", ">"),
        "operating_profit_margin_pct": ("opm_min", ">"),
        "pe_ratio": ("pe_max", "<"),
        "pb_ratio": ("pb_max", "<"),
        "dividend_yield_pct": ("dividend_yield_min", ">"),
        "dividend_payout_ratio_pct": ("dividend_payout_max", "<"),
        "interest_coverage": ("icr_min", ">="),
        "market_cap_crore": ("market_cap_min", ">"),
        "net_profit": ("net_profit_min", ">"),
        "eps_cagr_5yr": ("eps_cagr_min", ">"),
        "asset_turnover": ("asset_turnover_min", ">"),
        "sales": ("sales_min", ">"),
    }
    if col not in checks:
        return None
    rule, op = checks[col]
    threshold = rules.get(rule)
    if threshold is None:
        return None
    if col == "debt_to_equity" and row.get("broad_sector") == "Financials":
        return True
    if op == ">":
        return value > threshold
    if op == "<":
        return value < threshold
    if op == "<=":
        return value <= threshold
    if op == ">=":
        return value >= threshold
    return None


def build_screener_workbook():
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    all_presets = presets()

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        for name, rules in all_presets.items():
            df = run_preset(name).copy()
            df = df.sort_values("composite_quality_score", ascending=False)
            cols = [c for c in EXPORT_COLUMNS if c in df.columns]
            df[cols].to_excel(writer, sheet_name=_sheet_name(name), index=False)

    wb = load_workbook(OUTPUT_FILE)
    for preset_name, rules in all_presets.items():
        ws = wb[_sheet_name(preset_name)]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = HEADER
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=THIN)

        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2):
            row_data = {headers[i]: row[i].value for i in range(len(headers))}
            for idx, cell in enumerate(row):
                col = headers[idx]
                status = _meets(col, cell.value, rules, row_data)
                if status is True:
                    cell.fill = GREEN
                elif status is False:
                    cell.fill = RED
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"
                cell.border = Border(bottom=THIN)

        for col_cells in ws.columns:
            letter = col_cells[0].column_letter
            width = max(12, min(28, max(len(str(c.value)) if c.value is not None else 0 for c in col_cells) + 2))
            ws.column_dimensions[letter].width = width

    wb.save(OUTPUT_FILE)
    print(f"saved {OUTPUT_FILE}")


if __name__ == "__main__":
    build_screener_workbook()
