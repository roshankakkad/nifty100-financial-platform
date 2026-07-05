from pathlib import Path
import sqlite3

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

ROOT = Path(__file__).resolve().parents[2]
DB_PATH = ROOT / "data" / "db" / "nifty100.db"
OUT_FILE = ROOT / "output" / "peer_comparison.xlsx"

METRICS = [
    "return_on_equity_pct", "return_on_capital_employed_pct", "net_profit_margin_pct",
    "debt_to_equity", "free_cash_flow_cr", "pat_cagr_5yr", "revenue_cagr_5yr",
    "eps_cagr_5yr", "interest_coverage", "asset_turnover", "composite_quality_score"
]

GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
RED = PatternFill("solid", fgColor="F4CCCC")
GOLD = PatternFill("solid", fgColor="FFD966")
HEADER = PatternFill("solid", fgColor="385723")
THIN = Side(style="thin", color="D9EAD3")


def latest_ratios(conn):
    ratios = pd.read_sql("SELECT * FROM financial_ratios", conn)
    companies = pd.read_sql("SELECT id, company_name FROM companies", conn)
    ratios = ratios.merge(companies, left_on="company_id", right_on="id", how="left")
    ratios["year_num"] = pd.to_numeric(ratios["year_num"], errors="coerce")
    ratios = ratios.dropna(subset=["year_num", "return_on_equity_pct", "net_profit_margin_pct"])
    return ratios.sort_values(["company_id", "year_num"]).groupby("company_id", as_index=False).tail(1)


def source_tables():
    conn = sqlite3.connect(DB_PATH)
    ratios = latest_ratios(conn)
    peers = pd.read_sql("SELECT peer_group_name, company_id, is_benchmark FROM peer_groups", conn)
    pct = pd.read_sql("SELECT * FROM peer_percentiles", conn)
    conn.close()
    return ratios, peers, pct


def add_percentiles(data, pct):
    out = data.copy()
    for metric in METRICS:
        p = pct[pct["metric"] == metric][["company_id", "peer_group_name", "percentile_rank"]]
        p = p.rename(columns={"percentile_rank": metric + "_pct"})
        out = out.merge(p, on=["company_id", "peer_group_name"], how="left")
    return out


def build_peer_report():
    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    ratios, peers, pct = source_tables()

    with pd.ExcelWriter(OUT_FILE, engine="openpyxl") as writer:
        for group_name, members in peers.groupby("peer_group_name"):
            data = members.merge(ratios, on="company_id", how="left")
            data = add_percentiles(data, pct)
            cols = ["company_id", "company_name", "is_benchmark"] + [c for c in METRICS if c in data.columns] + [c + "_pct" for c in METRICS if c + "_pct" in data.columns]
            data = data[cols]
            med = {c: data[c].median(numeric_only=True) if pd.api.types.is_numeric_dtype(data[c]) else "" for c in data.columns}
            med["company_id"] = "PEER MEDIAN"
            data = pd.concat([data, pd.DataFrame([med])], ignore_index=True)
            sheet = group_name[:31]
            data.to_excel(writer, sheet_name=sheet, index=False)

    wb = load_workbook(OUT_FILE)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        headers = [c.value for c in ws[1]]
        for cell in ws[1]:
            cell.fill = HEADER
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(bottom=THIN)

        pct_cols = [i + 1 for i, h in enumerate(headers) if str(h).endswith("_pct")]
        bench_col = headers.index("is_benchmark") + 1 if "is_benchmark" in headers else None
        for row in ws.iter_rows(min_row=2):
            is_bench = bench_col and row[bench_col - 1].value == 1
            is_median = row[0].value == "PEER MEDIAN"
            for cell in row:
                cell.border = Border(bottom=THIN)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"
                if is_bench:
                    cell.fill = GOLD
                if is_median:
                    cell.font = Font(bold=True)
            if not is_bench and not is_median:
                for col_idx in pct_cols:
                    cell = row[col_idx - 1]
                    if cell.value is None:
                        continue
                    if cell.value >= 0.75:
                        cell.fill = GREEN
                    elif cell.value <= 0.25:
                        cell.fill = RED
                    else:
                        cell.fill = YELLOW
        for col_cells in ws.columns:
            letter = col_cells[0].column_letter
            width = max(12, min(30, max(len(str(c.value)) if c.value is not None else 0 for c in col_cells) + 2))
            ws.column_dimensions[letter].width = width

    wb.save(OUT_FILE)
    print(f"saved {OUT_FILE}")


if __name__ == "__main__":
    build_peer_report()
