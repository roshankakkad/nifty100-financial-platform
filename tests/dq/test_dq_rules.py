from pathlib import Path
import sqlite3

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
DB_PATH = ROOT / "data" / "db" / "nifty100.db"


def query(sql):
    conn = sqlite3.connect(DB_PATH)
    out = pd.read_sql(sql, conn)
    conn.close()
    return out


def test_companies_loaded():
    assert query("SELECT COUNT(*) c FROM companies")["c"].iloc[0] >= 90


def test_core_tables_loaded():
    for table in ["profitandloss", "balancesheet", "cashflow"]:
        assert query(f"SELECT COUNT(*) c FROM {table}")["c"].iloc[0] > 1000


def test_financial_ratios_row_count():
    assert query("SELECT COUNT(*) c FROM financial_ratios")["c"].iloc[0] >= 1100


def test_financial_ratios_has_required_columns():
    cols = set(query("SELECT * FROM financial_ratios LIMIT 1").columns)
    required = {
        "net_profit_margin_pct", "operating_profit_margin_pct", "return_on_equity_pct",
        "debt_to_equity", "interest_coverage", "asset_turnover", "free_cash_flow_cr",
        "capex_cr", "earnings_per_share", "book_value_per_share", "dividend_payout_ratio_pct",
        "total_debt_cr", "cash_from_operations_cr", "revenue_cagr_5yr", "pat_cagr_5yr",
        "eps_cagr_5yr", "composite_quality_score",
    }
    assert required.issubset(cols)


def test_no_null_only_required_columns():
    df = query("SELECT * FROM financial_ratios")
    required = ["return_on_equity_pct", "debt_to_equity", "free_cash_flow_cr", "revenue_cagr_5yr"]
    for col in required:
        assert df[col].notna().sum() > 0


def test_screener_output_exists_with_six_sheets():
    wb = load_workbook(ROOT / "output" / "screener_output.xlsx", read_only=True)
    assert len(wb.sheetnames) == 6


def test_peer_comparison_has_eleven_sheets():
    wb = load_workbook(ROOT / "output" / "peer_comparison.xlsx", read_only=True)
    assert len(wb.sheetnames) == 11


def test_radar_charts_created():
    charts = list((ROOT / "reports" / "radar_charts").glob("*_radar.png"))
    assert len(charts) >= 40


def test_capital_allocation_exists():
    assert (ROOT / "output" / "capital_allocation.csv").exists()


def test_ratio_edge_log_exists():
    assert (ROOT / "output" / "ratio_edge_cases.log").exists()


def test_screener_config_exists():
    assert (ROOT / "config" / "screener_config.yaml").exists()


def test_peer_percentiles_have_all_groups():
    assert query("SELECT COUNT(DISTINCT peer_group_name) c FROM peer_percentiles")["c"].iloc[0] == 11


def test_market_cap_loaded():
    assert query("SELECT COUNT(*) c FROM market_cap")["c"].iloc[0] >= 500


def test_stock_prices_loaded():
    assert query("SELECT COUNT(*) c FROM stock_prices")["c"].iloc[0] >= 5000
